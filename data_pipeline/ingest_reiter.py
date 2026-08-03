#!/usr/bin/env python3
"""
ingest_reiter.py — Ingest Andy Reiter's US News ranking-history datasets.

Source page: https://www.andyreiter.com/datasets/
Reiter publishes long-run US News ranking histories. We use two:

  national_liberal_arts:
    US-News-Rankings-Liberal-Arts-Colleges-Through-2026.xlsx
    (National Liberal Arts Colleges, ~1984-2026)
  national_universities:
    US-News-National-University-Rankings-Top-150-Through-2026.xlsx
    (National Universities top 150, ~1984-2026)

Both files are wide-format: one row per school, with the school's IPEDS UNITID
in a dedicated column and one column per ranking year. We melt them to long
format. Both files DO carry IPEDS UNITID (verified), so `unitid` is populated.

NOTE: Reiter does NOT cover the 8 US News regional categories — that is expected
and out of scope here; this script handles National Universities + National
Liberal Arts Colleges only.

Output: data/usn_rank_history.csv
Columns: unitid, name, year, rank, category
  category in {national_universities, national_liberal_arts}
"""

import csv
import os
import re
import sys
import urllib.request

import openpyxl

# Resolved from the datasets page (fetched 2026-06; links live under
# /wp-content/uploads/2025/09/). If these 404, re-fetch the page and relocate.
SOURCES = [
    {
        "category": "national_liberal_arts",
        "url": "https://andyreiter.com/wp-content/uploads/2025/09/"
               "US-News-Rankings-Liberal-Arts-Colleges-Through-2026.xlsx",
        "filename": "US-News-Rankings-Liberal-Arts-Colleges-Through-2026.xlsx",
    },
    {
        "category": "national_universities",
        "url": "https://andyreiter.com/wp-content/uploads/2025/09/"
               "US-News-National-University-Rankings-Top-150-Through-2026.xlsx",
        "filename": "US-News-National-University-Rankings-Top-150-Through-2026.xlsx",
    },
]

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_PATH = os.path.join(REPO, "data", "usn_rank_history.csv")
RAW_DIR = os.path.join(REPO, "data", "raw")


def download(url, dest):
    print(f"  downloading {url}", file=sys.stderr)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 PIPO-Index/1.0"})
    with urllib.request.urlopen(req, timeout=120) as resp:
        data = resp.read()
    with open(dest, "wb") as fh:
        fh.write(data)
    print(f"  saved {len(data)} bytes -> {dest}", file=sys.stderr)


def find_header_row(ws, max_scan=10):
    """Locate the header row (the one naming the columns).

    Files differ: the LAC sheet has a usage-note in row 1 and the real header
    in row 2; the National sheet's header is row 1. Detect by finding the row
    that contains an IPEDS column and at least one 4-digit year.
    """
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cells = [str(c).strip() if c is not None else "" for c in row]
        joined = " ".join(cells).lower()
        has_ipeds = "ipeds" in joined
        has_year = any(re.fullmatch(r"(19|20)\d{2}", c) for c in cells)
        if has_ipeds and has_year:
            return idx, cells
    raise RuntimeError("Could not locate header row (no IPEDS + year row found)")


def col_indexes(header):
    """Return (name_idx, ipeds_idx, {year:col_idx}) from a header row list."""
    name_idx = ipeds_idx = None
    year_cols = {}
    for i, h in enumerate(header):
        hl = h.lower()
        if name_idx is None and ("name" in hl):
            name_idx = i
        if ipeds_idx is None and ("ipeds" in hl):
            ipeds_idx = i
        m = re.fullmatch(r"(19|20)\d{2}", h)
        if m:
            year_cols[int(h)] = i
    if name_idx is None or ipeds_idx is None or not year_cols:
        raise RuntimeError(f"Header missing required columns: name={name_idx} ipeds={ipeds_idx} years={len(year_cols)}")
    return name_idx, ipeds_idx, year_cols


def parse_rank(value):
    """Return an int rank, or None for blanks / 'Rank Not Published' style cells."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # Some cells carry a tie marker or range; keep the leading integer.
    m = re.match(r"(\d+)", s)
    return int(m.group(1)) if m else None


def parse_file(source):
    path = os.path.join(RAW_DIR, source["filename"])
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row_idx, header = find_header_row(ws)
    name_idx, ipeds_idx, year_cols = col_indexes(header)

    out = []
    has_unitid_any = False
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if row is None:
            continue
        name = row[name_idx] if name_idx < len(row) else None
        if name is None or not str(name).strip():
            continue
        name = str(name).strip()
        raw_id = row[ipeds_idx] if ipeds_idx < len(row) else None
        unitid = ""
        if raw_id is not None:
            m = re.search(r"\d+", str(raw_id))
            if m:
                unitid = m.group(0)
                has_unitid_any = True
        for year, col in year_cols.items():
            rank = parse_rank(row[col]) if col < len(row) else None
            if rank is None:
                continue
            out.append({
                "unitid": unitid,
                "name": name,
                "year": year,
                "rank": rank,
                "category": source["category"],
            })
    return out, has_unitid_any, sorted(year_cols)


def main():
    os.makedirs(RAW_DIR, exist_ok=True)
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)

    all_rows = []
    report = []
    for source in SOURCES:
        dest = os.path.join(RAW_DIR, source["filename"])
        download(source["url"], dest)
        rows, has_unitid, years = parse_file(source)
        all_rows.extend(rows)
        report.append((source["category"], len(rows), has_unitid, min(years), max(years)))
        print(f"  {source['category']}: {len(rows)} rows, "
              f"unitid={'YES' if has_unitid else 'NO'}, years {min(years)}-{max(years)}",
              file=sys.stderr)

    fields = ["unitid", "name", "year", "rank", "category"]
    with open(OUT_PATH, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"Wrote {len(all_rows)} rows -> {OUT_PATH}")
    print("UNITID coverage by category:")
    for cat, n, has_id, ymin, ymax in report:
        print(f"  {cat}: {n} rows, UNITID present={has_id}, years {ymin}-{ymax}")
    return all_rows


if __name__ == "__main__":
    main()
