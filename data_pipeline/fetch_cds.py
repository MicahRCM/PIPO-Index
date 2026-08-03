#!/usr/bin/env python3
"""
fetch_cds.py — Production Common Data Set (CDS) extraction pipeline for PIPO-Index.

FREE TIER (no browser): downloads CDS files with curl and parses them with
pdfplumber (PDF) or a tag-stripping text parser (HTML). One parser keyed on the
standardized CDS section codes (C9, C12, F1, H2) so it generalizes across layouts.

Extracts the six CDS-only variables:
  hs_gpa          (C12)  Average high school GPA
  pct_top10_hs    (C9)   Percent in top tenth of HS graduating class
  pct_need_met    (H2)   Average percentage of need met
  need_aid_total  (H2)   Total need-based scholarship/grant dollars
  merit_aid_total (H2)   Total non-need-based (merit) scholarship/grant dollars
  avg_aid_package (H2 J) Average financial aid package
  pct_commuter    (F1)   Percent living off campus / commuting
  pct_demonstrated_need  (H2 lines A & C)  Share of the entering class with
                         demonstrated financial need = 100 * (line C: first-time
                         first-year students determined to have financial need) /
                         (line A: first-time first-year degree-seeking students
                         enrolled). The first-time first-year column is used so
                         it lines up with pct_need_met / avg_aid_package. Using
                         enrolled (line A) as the denominator — rather than line
                         B (applied) — gives the fraction of the whole class that
                         demonstrated need, the more interpretable school metric.

Anti-traps (from the proving run):
  * Every "Common Data Set 20XX-20XX" header is verified INSIDE the file, and each
    extracted value is tagged with the year of its NEAREST preceding header. Values
    whose section year differs from the chosen cds_year are dropped (the Amherst
    split-PDF mismatch, e.g. C=2024-25 but F/H=2019-20).
  * Cloudflare "Just a moment" / HTTP 403, Box, and Tableau JS apps are NOT chased;
    they are logged to cds_blocked.csv with a reason for the later browser tier.

Outputs (checkpointed after every school, fully resumable):
  data/cds_data.csv     extracted values
  data/cds_blocked.csv  schools that need the browser tier

USAGE
  Batch (reads data/cds_discovered_urls.csv: unitid,url[,name]):
      python3 data_pipeline/fetch_cds.py --batch
  Single URL:
      python3 data_pipeline/fetch_cds.py --one --unitid 139818 \
          --name "Agnes Scott College" --url <cds_url>
  Parse a local file (testing):
      python3 data_pipeline/fetch_cds.py --parse-file path.pdf [--url hint]
"""
from __future__ import annotations

import argparse
import contextlib
import csv
import io
import json
import os
import re
import signal
import subprocess
import sys
from typing import Optional

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA = os.path.join(ROOT, "data")
CACHE = os.path.join(DATA, "cds_cache")
UNIVERSITY_JSON = os.path.join(ROOT, "web", "data", "university_data.json")
IPEDS_DIR = os.path.join(DATA, "ipeds_directory.csv")
DISCOVERED = os.path.join(DATA, "cds_discovered_urls.csv")
OUT_DATA = os.path.join(DATA, "cds_data.csv")
OUT_BLOCKED = os.path.join(DATA, "cds_blocked.csv")

DATA_FIELDS = [
    "unitid", "name", "cds_year", "cds_url",
    "hs_gpa", "pct_top10_hs", "pct_need_met",
    "merit_aid_total", "need_aid_total", "avg_aid_package", "pct_commuter",
    "pct_demonstrated_need",
]
BLOCKED_FIELDS = ["unitid", "name", "attempted_url", "reason"]

ACCEPTABLE_YEARS = ("2025-2026", "2024-2025", "2023-2024")
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 " \
     "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"


# ---------------------------------------------------------------------------
# Year handling
# ---------------------------------------------------------------------------
_HEADER_RE = re.compile(
    r"common data set\s*(20\d\d)\s*[-‐-―/]\s*((?:20)?\d\d)", re.I)


def normalize_year(start: str, end: str) -> str:
    """('2024','25') or ('2024','2025') -> '2024-2025'."""
    end = end.strip()
    if len(end) == 2:
        end = start[:2] + end
    return f"{start}-{end}"


def find_headers(text: str):
    """Return list of (position, 'YYYY-YYYY') for every CDS header in the file."""
    out = []
    for m in _HEADER_RE.finditer(text):
        out.append((m.start(), normalize_year(m.group(1), m.group(2))))
    return out


def nearest_year(headers, pos: int) -> Optional[str]:
    """Year of the header immediately preceding `pos` (the section a value sits in)."""
    best = None
    for hpos, yr in headers:
        if hpos <= pos:
            best = yr
        else:
            break
    if best is None and headers:
        best = headers[0][1]  # value appears before first header on the page
    return best


# ---------------------------------------------------------------------------
# Value extraction helpers
# ---------------------------------------------------------------------------
_PCT = r"(\d{1,3}(?:\.\d+)?)\s*%"


def _clean_gpa(raw) -> Optional[float]:
    if raw is None:
        return None
    raw = re.sub(r"\.{2,}", ".", str(raw)).strip(". ")  # OCR artifact "3..71" -> "3.71"
    m = re.match(r"^([0-5](?:\.\d{1,2})?)", raw)
    if not m:
        return None
    val = float(m.group(1))
    return val if 0.0 < val <= 5.0 else None


def _to_int(s: str) -> int:
    return int(s.replace(",", ""))


def _norm_pct(raw, frac_below: float = 1.0) -> Optional[float]:
    """Coerce a raw cell/field value to a 0-100 percentage.

    `frac_below` controls fraction handling: a value v with 0 < v < frac_below is
    interpreted as a fraction and multiplied by 100. Use 1.5 for xlsx percent
    cells (which store 1.0 == 100%); use 1.0 for AcroForm fields (which store
    whole-number percents, so only sub-1 typos like '.75' get rescaled).
    """
    if raw is None:
        return None
    s = str(raw).strip().replace("%", "").replace(",", "")
    if s in ("", "None"):
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    v = float(m.group())
    if 0 < v < frac_below:
        v *= 100.0
    if 0 <= v <= 100:
        return round(v, 2)
    return None


def _norm_money(raw) -> Optional[int]:
    """Coerce a raw cell/field value (e.g. '$1,234', 1234.5) to an int dollar amount."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s in ("", "None"):
        return None
    m = re.search(r"\d[\d,]*(?:\.\d+)?", s)
    if not m:
        return None
    return int(round(float(m.group().replace(",", ""))))


def _dollar_amounts(window: str, n: int):
    """
    Extract up to `n` dollar amounts from a window of CDS text.
    Columns are '$'-delimited; within a column the value is at the start of the
    line. Strips comma separators AND stray spaces that pdfplumber injects inside
    numbers (e.g. '$ 1 48,278,012' -> 148278012).
    """
    out = []
    for seg in window.split("$")[1:]:
        seg = seg.split("\n")[0]                  # same line only
        cleaned = seg.replace(",", "").replace(" ", "")
        m = re.match(r"(\d+)", cleaned)
        if m:
            out.append(int(m.group(1)))
        if len(out) >= n:
            break
    return out


# A CDS H2 data row is 2 or 3 whitespace-separated integers (the First-time
# First-year / Full-time Undergrad / Less-than-full-time columns). The leading
# negative lookbehind keeps us from starting in the middle of a number, so
# stray label digits like "B1" or "Fall 2024" don't match (they aren't followed
# by a second integer).
_H2_ROW = re.compile(
    r"(?<![\d.,])(\d[\d,]*)\s+(\d[\d,]*)(?:\s+(\d[\d,]*))?(?![\d.,])")


def _h2_last_col(segment: str):
    """First-column integer of the LAST H2 data row in `segment`, or None."""
    last = None
    for m in _H2_ROW.finditer(segment):
        last = m
    return _to_int(last.group(1)) if last else None


def _demonstrated_need_from_text(text: str):
    """Compute pct_demonstrated_need from CDS H2 lines A and C (first-time
    first-year column). Returns (pct, anchor_pos) or (None, None).

    Everything is anchored on line B ("...applied for need-based financial aid",
    the phrase unique to H2), because a data row can wrap to EITHER side of its
    label depending on the PDF: some schools print the numbers after the label
    words, others in the middle. Relative to line B the rows are unambiguous:
      * line A (enrolled)  = the last data row in the text just BEFORE line B
      * line C (has need)  = the 2nd data row at/after line B (row 1 is line B's)
    % = 100 * C / A — the share of the entering class with demonstrated need.
    """
    low = text.lower()
    # Tolerate a data row printed between "applied for" and "need-based"
    # (Michigan wraps line B's numbers into the middle of its own label).
    mb = re.search(r"applied for[\s\d,]*need[- ]?based", low)
    if not mb:
        return None, None
    bpos = mb.start()
    a = _h2_last_col(text[max(0, bpos - 400):bpos])          # line A (enrolled)
    rows = list(_H2_ROW.finditer(text[bpos:bpos + 400]))     # [B, C, D, ...]
    c = _to_int(rows[1].group(1)) if len(rows) >= 2 else None
    if a and c and a > 0 and 0 <= c <= a:
        return round(100.0 * c / a, 2), bpos
    return None, None


def extract_values(text: str):
    """
    Run the section-coded parser over CDS plain text.
    Returns dict: {field: (value, char_position)} for every field found.
    Positions let the caller tag each value with its section's CDS year.
    """
    found = {}
    low = text.lower()

    # --- C12: Average high school GPA -------------------------------------
    m = re.search(r"submitted gpa[:\s]*\$?\s*([0-5][\d.]*)", low)
    if m:
        g = _clean_gpa(m.group(1))
        if g is not None:
            found["hs_gpa"] = (g, m.start())

    # --- C9: Percent in top tenth of HS class (inline % only) -------------
    m = re.search(r"top tenth of high school graduating class[:\s]*" + _PCT, low)
    if m:
        found["pct_top10_hs"] = (float(m.group(1)), m.start())

    # --- F1: Percent living off campus or commuting (last % on line) ------
    m = re.search(r"live off campus or commute([^\n]*)", low)
    if m:
        pcts = re.findall(_PCT, m.group(1))
        if pcts:
            found["pct_commuter"] = (float(pcts[-1]), m.start())

    # --- H2: Average percent of need met (first number after anchor) ------
    m = re.search(r"percentage of need that was met", low)
    if m:
        window = text[m.end():m.end() + 320]
        pm = re.search(_PCT, window)
        if pm:
            found["pct_need_met"] = (float(pm.group(1)), m.start())

    # --- H2: Total Scholarships/Grants -> need total, merit total ---------
    m = re.search(r"total scholarships\s*/\s*grants", low)
    if m:
        amts = _dollar_amounts(text[m.end():m.end() + 200], 2)
        if len(amts) >= 1:
            found["need_aid_total"] = (amts[0], m.start())
        if len(amts) >= 2:
            found["merit_aid_total"] = (amts[1], m.start())

    # --- H2 line J: Average financial aid package (first $ after anchor) ---
    m = re.search(r"average financial aid package", low)
    if m:
        amts = _dollar_amounts(text[m.end():m.end() + 320], 1)
        if amts:
            found["avg_aid_package"] = (amts[0], m.start())

    # --- H2 lines A & C: percent of the class with demonstrated need ---
    dn, dpos = _demonstrated_need_from_text(text)
    if dn is not None:
        found["pct_demonstrated_need"] = (dn, dpos)

    return found


def parse_file_text(text: str, url_hint: str = ""):
    """
    Parse the text of ONE CDS file.
    Returns (values, file_year) where:
      values    = {field: (value, section_year)}  section_year = nearest header
      file_year = the file's academic CDS year (year nearest GPA/top-10, else
                  most recent header, else year guessed from the URL)
    """
    headers = find_headers(text)
    raw = extract_values(text)

    anchor_pos = None
    for key in ("hs_gpa", "pct_top10_hs"):
        if key in raw:
            anchor_pos = raw[key][1]
            break
    if headers:
        file_year = nearest_year(headers, anchor_pos) if anchor_pos is not None \
            else max(y for _, y in headers)
    else:
        file_year = _year_from_url(url_hint)

    values = {}
    for field, (val, pos) in raw.items():
        sect_year = nearest_year(headers, pos) if headers else file_year
        values[field] = (val, sect_year)
    return values, file_year


def parse_cds_text(text: str, url_hint: str = ""):
    """
    Parse a single-file CDS into a result row + diagnostics.
    Returns (row_dict_or_None, status, reason)  status: 'ok'|'stale_only'|'parse_fail'
    """
    values, file_year = parse_file_text(text, url_hint)
    if not values:
        return None, "parse_fail", "pdf_parse_fail"
    return _assemble_row([(values, file_year)])


def _assemble_row(parsed_files):
    """
    Merge one or more parsed CDS files into a single output row.
    parsed_files: list of (values, file_year).
    Picks cds_year from the file carrying the academic (C) section, else the most
    recent acceptable file year. Drops any value whose own section year != cds_year
    (defends against split files with mismatched years — the Amherst trap).
    """
    # cds_year: prefer the file that yielded GPA / top-10 (the academic section)
    cds_year = None
    for values, fy in parsed_files:
        if ("hs_gpa" in values or "pct_top10_hs" in values) and fy in ACCEPTABLE_YEARS:
            cds_year = fy
            break
    if cds_year is None:
        acceptable = [fy for _, fy in parsed_files if fy in ACCEPTABLE_YEARS]
        if acceptable:
            cds_year = max(acceptable)

    if cds_year is None:
        return None, "stale_only", "stale_only"

    row = {f: "" for f in DATA_FIELDS}
    row["cds_year"] = cds_year
    for values, _ in parsed_files:
        for field, (val, sect_year) in values.items():
            if row[field] != "":
                continue
            if sect_year and sect_year != cds_year:
                continue  # stale-year section, skip
            row[field] = val

    if not any(row[f] != "" for f in DATA_FIELDS[4:]):
        return None, "stale_only", "stale_only"
    return row, "ok", ""


def _year_from_url(url: str) -> Optional[str]:
    if not url:
        return None
    # Prefer a *consecutive* year pair (a real CDS year like 2024-2025) over a
    # spurious date-folder match (e.g. ".../2025-03/..." -> 2025-2003).
    best = None
    for m in re.finditer(r"(20\d\d)[-_](20\d\d|\d\d)", url):
        yr = normalize_year(m.group(1), m.group(2))
        start, end = yr.split("-")
        if end.isdigit() and int(end) == int(start) + 1:
            return yr
        if best is None:
            best = yr
    return best


# ---------------------------------------------------------------------------
# Fetch (curl) + format detection
# ---------------------------------------------------------------------------
def _blocked_by_url(url: str) -> Optional[str]:
    u = url.lower()
    if "box.com" in u:
        return "box"
    if "tableau" in u or "public.tableau" in u:
        return "tableau"
    return None


def fetch(url: str, dest: str):
    """
    Download `url` to `dest` with curl.
    Returns (kind, reason) where kind in {'pdf','html','blocked','error'}.
    """
    pre = _blocked_by_url(url)
    if pre:
        return "blocked", pre

    try:
        proc = subprocess.run(
            ["curl", "-sL", "--max-time", "45", "-A", UA,
             "-w", "%{http_code}\t%{content_type}", "-o", dest, url],
            capture_output=True, text=True, timeout=60,
        )
    except subprocess.TimeoutExpired:
        return "error", "not_found"
    if proc.returncode != 0:
        return "error", "not_found"

    meta = (proc.stdout or "").strip().split("\t")
    http_code = meta[0] if meta else ""
    ctype = meta[1].lower() if len(meta) > 1 else ""

    if not os.path.exists(dest) or os.path.getsize(dest) == 0:
        return "error", "not_found"

    with open(dest, "rb") as fh:
        head = fh.read(4096)

    # Cloudflare interstitial / challenge
    low = head.lower()
    if b"just a moment" in low or b"cf-browser-verification" in low \
            or b"challenge-platform" in low or http_code == "403":
        return "blocked", "cloudflare"

    if head[:5] == b"%PDF-" or "application/pdf" in ctype:
        return "pdf", ""

    # .xlsx (and other OOXML) files are ZIP archives starting with 'PK'.
    if head[:2] == b"PK" and (
            ".xlsx" in url.lower() or "spreadsheet" in ctype
            or "officedocument" in ctype or b"[Content_Types].xml" in head):
        return "xlsx", ""

    if b"box.com" in low or b"boxcdn" in low:
        return "blocked", "box"
    if b"tableau" in low:
        return "blocked", "tableau"

    if http_code and http_code not in ("200", "206"):
        return "error", "not_found"

    return "html", ""


def _find_cds_link(html: str, base_url: str) -> Optional[str]:
    """Find a direct CDS .pdf/.xlsx link inside a landing-page's HTML.

    Many schools publish an HTML index that links to the actual CDS file. We
    pick the newest same-CDS file link (href contains 'cds' or 'common'+'data'),
    preferring an acceptable-year URL. Returns an absolute URL or None. Free tier:
    this is one extra curl hop, not a browser.
    """
    from urllib.parse import urljoin
    cands = []
    for m in re.finditer(r'href\s*=\s*["\']([^"\']+\.(?:pdf|xlsx))["\']', html, re.I):
        href = m.group(1)
        low = href.lower()
        if "cds" not in low and not ("common" in low and "data" in low):
            continue
        yr = _year_from_url(href) or ""
        cands.append((yr, urljoin(base_url, href)))
    if not cands:
        return None
    cands.sort(key=lambda t: t[0], reverse=True)          # newest year first
    for yr, url in cands:
        if yr in ACCEPTABLE_YEARS:
            return url
    return cands[0][1]


def pdf_to_text(path: str, max_pages: int = 60) -> str:
    """Extract text from a CDS PDF. Capped at `max_pages` because the CDS C/F/H
    sections we parse sit in the first ~35 pages; the cap bounds cost on the
    occasional multi-megabyte / image-heavy PDF."""
    import pdfplumber
    parts = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages[:max_pages]:
            parts.append(page.extract_text() or "")
    return "\n".join(parts)


def html_to_text(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as fh:
        html = fh.read()
    html = re.sub(r"(?is)<(script|style).*?</\1>", " ", html)
    html = re.sub(r"(?s)<[^>]+>", " ", html)
    html = (html.replace("&amp;", "&").replace("&nbsp;", " ")
            .replace("&#39;", "'").replace("&quot;", '"'))
    return re.sub(r"[ \t]+", " ", html)


# ---------------------------------------------------------------------------
# Fallback 1: AcroForm / widget-annotation reader (pypdf)
#
# Many universities publish the official fillable CDS template. pdfplumber's
# extract_text() returns the static labels but NOT the typed-in values (they live
# in form-field / widget annotations). pypdf reads those. The template uses a
# stable set of field names; the map below pairs each of our six variables with
# its standard field. need_aid_total / merit_aid_total come from the need-based
# vs non-need (merit) total scholarship/grant fields; the percent/package fields
# use the first-time-freshman column, matching the line-oriented text parser.
# ---------------------------------------------------------------------------
ACROFORM_MAP = {
    "hs_gpa":          ("FRSH_GPA",          _clean_gpa),
    "pct_top10_hs":    ("FRSH_HS_RANK_10_P", lambda v: _norm_pct(v, 1.0)),
    "pct_need_met":    ("FRSH_FT_ND_MET_P",  lambda v: _norm_pct(v, 1.0)),
    "need_aid_total":  ("SCHOL_NB_TOT_D",    _norm_money),
    "merit_aid_total": ("SCHOL_NN_TOT_D",    _norm_money),
    "avg_aid_package": ("FRSH_FT_AVG_PKG_D", _norm_money),
    "pct_commuter":    ("HOUS_COMMUTE_P",    lambda v: _norm_pct(v, 1.0)),
}


def _acroform_raw_fields(path: str):
    """Return {field_name: raw_value} for a fillable PDF, or {} if it has none.

    Tries the AcroForm field tree first, then falls back to scanning page widget
    annotations (some flattened/partial forms expose values only there).
    """
    try:
        import pypdf
    except ImportError:  # pragma: no cover
        return {}
    try:
        reader = pypdf.PdfReader(path)
        fields = reader.get_fields() or {}
    except Exception:
        return {}

    out = {}
    for k, v in fields.items():
        try:
            out[str(k)] = v.get("/V")
        except Exception:
            continue
    if out:
        return out

    # No field tree — scan widget annotations directly.
    try:
        for page in reader.pages:
            for a in page.get("/Annots") or []:
                o = a.get_object()
                if o.get("/Subtype") == "/Widget" and o.get("/T") is not None:
                    out[str(o.get("/T"))] = o.get("/V")
    except Exception:
        pass
    return out


def acroform_values(path: str, text: str = "", url_hint: str = ""):
    """Extract the six CDS variables from a fillable-PDF's form fields.

    Returns (values, file_year) where values = {field: (val, section_year)};
    ({}, None) if the file carries no usable form fields.
    """
    raw = _acroform_raw_fields(path)
    if not raw:
        return {}, None

    # Year: prefer a verified header in the static text, then the ACAD_YR field,
    # then the URL.
    headers = find_headers(text) if text else []
    file_year = max((y for _, y in headers), default=None)
    if file_year is None:
        acad = raw.get("ACAD_YR")
        if acad:
            m = re.search(r"20\d\d", str(acad))
            if m:
                file_year = f"{m.group()}-{int(m.group()) + 1}"
    if file_year is None:
        file_year = _year_from_url(url_hint)

    values = {}
    for field, (fname, conv) in ACROFORM_MAP.items():
        if raw.get(fname) in (None, ""):
            continue
        val = conv(raw.get(fname))
        if val is not None and val != "":
            values[field] = (val, file_year)

    # pct_demonstrated_need is computed from two H2 count fields (no single field
    # holds it): FRSH_FT_N = line A (enrolled), FRSH_FT_ND_N = line C (determined
    # to have need), first-time first-year column.
    a = _norm_money(raw.get("FRSH_FT_N"))
    c = _norm_money(raw.get("FRSH_FT_ND_N"))
    if a and c and a > 0 and 0 <= c <= a:
        values["pct_demonstrated_need"] = (round(100.0 * c / a, 2), file_year)

    return values, file_year


# ---------------------------------------------------------------------------
# Fallback 2: spreadsheet (.xlsx) CDS — openpyxl
# Several schools publish the CDS as an Excel workbook with one sheet per
# section. Rows are [label, value1, value2, ...]; percents are stored as
# fractions (0.732 == 73.2%). We pair each label row with its numeric cells.
# ---------------------------------------------------------------------------
def _xlsx_row_values(cells, values):
    """Match one spreadsheet row against the CDS anchors and fill `values`."""
    label = " ".join(str(c) for c in cells if isinstance(c, str)).lower()
    nums = []
    for c in cells:
        if isinstance(c, bool):
            continue
        if isinstance(c, (int, float)):
            nums.append(float(c))
        elif isinstance(c, str) and re.fullmatch(r"\$?[\d,]+(?:\.\d+)?%?", c.strip()):
            nums.append(float(c.strip().lstrip("$").rstrip("%").replace(",", "")))

    def put(field, val):
        if val is not None and val != "" and field not in values:
            values[field] = val

    if "top tenth of high school graduating class" in label and nums:
        put("pct_top10_hs", _norm_pct(nums[0], 1.5))
    if "average high school gpa" in label:
        put("hs_gpa", _clean_gpa(str(cells[-1])))
    if "percentage of need that was met" in label and nums:
        put("pct_need_met", _norm_pct(nums[0], 1.5))
    if "total scholarships/grants" in label:
        if len(nums) >= 1:
            put("need_aid_total", _norm_money(nums[0]))
        if len(nums) >= 2:
            put("merit_aid_total", _norm_money(nums[1]))
    if "average financial aid package" in label and nums:
        put("avg_aid_package", _norm_money(nums[0]))
    if "live off campus or commute" in label and nums:
        put("pct_commuter", _norm_pct(nums[-1], 1.5))
    # H2 lines A (enrolled) and C (determined to have need), first-year column.
    # Stashed privately; xlsx_values divides them into pct_demonstrated_need.
    if "number of degree-seeking undergraduate students" in label and nums:
        put("_h2_a", nums[0])
    if "determined to have financial need" in label and nums:
        put("_h2_c", nums[0])


def xlsx_values(path: str, url_hint: str = ""):
    """Extract the six CDS variables from an .xlsx CDS workbook.

    Returns (values, file_year) with values = {field: (val, file_year)}.
    """
    try:
        import openpyxl
    except ImportError:  # pragma: no cover
        return {}, None
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(open(path, "rb").read()), read_only=True, data_only=True)
    except Exception:
        return {}, None

    raw = {}
    header_years = []
    bare_years = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(values_only=True):
            cells = [c for c in row if c is not None and str(c).strip() != ""]
            if not cells:
                continue
            joined = " ".join(str(c) for c in cells)
            for _, yr in find_headers(joined):
                header_years.append(yr)
            # bare "YYYY-YYYY" cells (CDS xlsx often omit the "Common Data Set" prefix)
            for m in re.finditer(r"(20\d\d)\s*[-‐-―/]\s*((?:20)?\d\d)", joined):
                bare_years.append(normalize_year(m.group(1), m.group(2)))
            _xlsx_row_values(cells, raw)

    a = raw.pop("_h2_a", None)
    c = raw.pop("_h2_c", None)
    if a and c and a > 0 and 0 <= c <= a:
        raw["pct_demonstrated_need"] = round(100.0 * c / a, 2)

    if header_years:
        file_year = max(header_years)
    else:
        # filename/url year is authoritative; bare cell years only as last resort
        # (a stray "2025-2026" projection column shouldn't override CDS_2024-2025)
        acceptable_bare = [y for y in bare_years if y in ACCEPTABLE_YEARS]
        file_year = _year_from_url(url_hint) or (max(acceptable_bare) if acceptable_bare else None)
    values = {f: (v, file_year) for f, v in raw.items()}
    return values, file_year


# ---------------------------------------------------------------------------
# Fallback 3: positional / two-column PDF layouts (pdfplumber word x-positions)
# Some CDS PDFs place the value in a right-hand column the line-oriented parser
# can't follow. We group words into visual rows, find each anchor label, then
# read numeric tokens sitting to its right (or just below in the value column).
# ---------------------------------------------------------------------------
_POS_ANCHORS = [
    ("pct_top10_hs",    r"top tenth of high school graduating class", "pct"),
    ("hs_gpa",          r"average high school gpa",                   "gpa"),
    ("pct_need_met",    r"percentage of need that was met",           "pct"),
    ("avg_aid_package", r"average financial aid package",             "money"),
    ("pct_commuter",    r"live off campus or commute",                "pct"),
]


def positional_values(path: str, text: str = "", url_hint: str = ""):
    """Pair labels to values by x/y position for two-column CDS PDFs.

    Returns (values, file_year). Best-effort: only emits a value when a numeric
    token sits on/right-of the label row (or the row just below it).
    """
    try:
        import pdfplumber
    except ImportError:  # pragma: no cover
        return {}, None

    headers = find_headers(text) if text else []
    file_year = max((y for _, y in headers), default=None) or _year_from_url(url_hint)

    # The labels we pair on must appear in the extracted text; if none do, the
    # positional pass cannot help — skip it (avoids costly word extraction on
    # large image/encoded PDFs that have no text-layer values).
    any_pat = re.compile("|".join(p for _, p, _ in _POS_ANCHORS))
    if text and not any_pat.search(text.lower()):
        return {}, file_year

    from collections import defaultdict
    values = {}
    n_fields = len(_POS_ANCHORS)
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages[:40]:  # cap pages; CDS sections are up front
                if len(values) >= n_fields:
                    break
                # Only word-extract pages that actually carry an anchor label
                # (cheap text scan first — extract_words is the expensive call).
                try:
                    ptext = (page.extract_text() or "").lower()
                except Exception:
                    continue
                if not any_pat.search(ptext):
                    continue
                try:
                    words = page.extract_words()
                except Exception:
                    continue
                rows = defaultdict(list)
                for w in words:
                    rows[round(w["top"])].append(w)
                ordered = sorted(rows)
                for ri, y in enumerate(ordered):
                    row = sorted(rows[y], key=lambda w: w["x0"])
                    line = " ".join(w["text"] for w in row).lower()
                    label_end = max((w["x1"] for w in row), default=0)
                    for field, pat, kind in _POS_ANCHORS:
                        if field in values or not re.search(pat, line):
                            continue
                        # candidate tokens: right of label on this row, then the
                        # row just below.
                        cands = [w for w in row if w["x0"] > label_end - 1
                                 and re.search(r"\d", w["text"])]
                        if not cands and ri + 1 < len(ordered):
                            below = rows[ordered[ri + 1]]
                            cands = [w for w in below if re.search(r"\d", w["text"])]
                        val = _coerce_token(cands, kind)
                        if val is not None:
                            values[field] = (val, file_year)
    except Exception:
        return values, file_year
    return values, file_year


def _coerce_token(cands, kind):
    if not cands:
        return None
    tokens = [w["text"] for w in cands]
    if kind == "gpa":
        for t in tokens:
            g = _clean_gpa(t)
            if g is not None:
                return g
        return None
    if kind == "money":
        for t in tokens:
            m = _norm_money(t)
            if m is not None and m >= 100:
                return m
        return None
    # pct: prefer a token bearing '%', else the last numeric token
    for t in tokens:
        if "%" in t:
            return _norm_pct(t, 1.0)
    return _norm_pct(tokens[-1], 1.0)


# ---------------------------------------------------------------------------
# Unified single-file parsers (text -> AcroForm -> positional)
# ---------------------------------------------------------------------------
def parse_pdf_file(path: str, url_hint: str = ""):
    """Parse one CDS PDF, returning (values, file_year).

    Runs the line-oriented text parser first; if it leaves variables unfilled,
    unions in AcroForm form-field values, then a positional pass. Text-extracted
    values always win on conflict.
    """
    text = pdf_to_text(path)
    values, file_year = parse_file_text(text, url_hint)

    if len(values) < len(DATA_FIELDS[4:]):
        av, ay = acroform_values(path, text, url_hint)
        for f, pair in av.items():
            values.setdefault(f, pair)
        file_year = file_year or ay
        if len(values) < len(DATA_FIELDS[4:]):
            pv, py = positional_values(path, text, url_hint)
            for f, pair in pv.items():
                values.setdefault(f, pair)
            file_year = file_year or py
    return values, file_year


# ---------------------------------------------------------------------------
# Roster + checkpoint
# ---------------------------------------------------------------------------
def load_national_schools():
    with open(UNIVERSITY_JSON) as fh:
        rows = json.load(fh)
    nat = [r for r in rows if r.get("nationalu") == 1]

    ipeds = {}
    with open(IPEDS_DIR, newline="") as fh:
        for r in csv.DictReader(fh):
            try:
                ipeds[int(r["unitid"])] = r
            except (ValueError, KeyError):
                pass

    schools = []
    for r in nat:
        uid = r.get("UNITID")
        info = ipeds.get(uid, {})
        schools.append({
            "unitid": uid,
            "name": r.get("Name", ""),
            "city": info.get("city", ""),
            "state": info.get("state", r.get("state", "")),
        })
    schools.sort(key=lambda s: s["unitid"])
    return schools


def _load_done():
    done = set()
    for path in (OUT_DATA, OUT_BLOCKED):
        if os.path.exists(path):
            with open(path, newline="") as fh:
                for r in csv.DictReader(fh):
                    if r.get("unitid"):
                        done.add(int(r["unitid"]))
    return done


def _append(path: str, fields, row):
    new = not os.path.exists(path)
    with open(path, "a", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        if new:
            w.writeheader()
        w.writerow(row)


def write_data(row):
    _append(OUT_DATA, DATA_FIELDS, row)


def write_blocked(unitid, name, url, reason):
    _append(OUT_BLOCKED, BLOCKED_FIELDS,
            {"unitid": unitid, "name": name, "attempted_url": url, "reason": reason})


# ---------------------------------------------------------------------------
# Process one school
# ---------------------------------------------------------------------------
def _parse_dest(dest: str, suburl: str):
    """Sniff a downloaded file's real type and parse it to (values, file_year).

    Type is determined from the magic bytes (not the URL), so cached `.bin`
    files and content-type mismatches still route correctly.
    """
    with open(dest, "rb") as fh:
        head = fh.read(8)
    if head[:5] == b"%PDF-":
        return parse_pdf_file(dest, suburl)
    if head[:2] == b"PK":
        return xlsx_values(dest, suburl)
    return parse_file_text(html_to_text(dest), suburl)


def _attempt(unitid, name, url, use_cache=False):
    """Fetch (or reuse cached) + parse a CDS for one school WITHOUT writing CSVs.

    `url` may be a single URL or several pipe-separated URLs (split-section
    schools). Returns (row_or_None, reason). When `use_cache` is True, any
    already-downloaded `data/cds_cache/{unitid}_*.bin` files are parsed first and
    a network fetch is only attempted if they yield nothing (idempotent re-runs).
    """
    os.makedirs(CACHE, exist_ok=True)
    if not url:
        return None, "not_found"

    suburls = [u.strip() for u in url.split("|") if u.strip()]
    parsed_files = []
    block_reason = None

    if use_cache:
        import glob as _glob
        cached = sorted(_glob.glob(os.path.join(CACHE, f"{unitid}_*.bin")))
        had_doc_cache = False  # a real PDF/xlsx was already downloaded
        for cb in cached:
            with open(cb, "rb") as fh:
                magic = fh.read(8)
            if magic[:5] == b"%PDF-" or magic[:2] == b"PK":
                had_doc_cache = True
            hint = suburls[0] if suburls else url
            try:
                values, file_year = _parse_dest(cb, hint)
            except Exception:
                continue
            if values:
                parsed_files.append((values, file_year))
        if parsed_files:
            row, status, breason = _assemble_row(parsed_files)
            if status == "ok" and row is not None:
                return row, ""
        # A cached PDF/xlsx is the same bytes the server would return, so a
        # re-fetch can't improve on it — skip the network round-trip (and the
        # costly second parse). Only HTML/error/missing caches warrant a retry.
        if had_doc_cache:
            return None, breason if parsed_files else "pdf_parse_fail"

    parsed_files = []
    for i, suburl in enumerate(suburls):
        dest = os.path.join(CACHE, f"{unitid}_{i}.bin")
        kind, reason = fetch(suburl, dest)
        if kind in ("blocked", "error"):
            block_reason = block_reason or reason
            continue
        try:
            values, file_year = _parse_dest(dest, suburl)
        except Exception:
            values, file_year = {}, None
        # HTML that yielded nothing may be a landing page linking to the real
        # CDS file — follow one hop to that file and parse it instead.
        if not values and kind == "html":
            values, file_year = _resolve_landing(dest, suburl, unitid, i)
        if values:
            parsed_files.append((values, file_year))
        elif not block_reason:
            block_reason = "pdf_parse_fail"

    if not parsed_files:
        return None, block_reason or "pdf_parse_fail"

    row, status, breason = _assemble_row(parsed_files)
    if status != "ok" or row is None:
        return None, breason or "pdf_parse_fail"
    return row, ""


def _resolve_landing(dest, suburl, unitid, idx):
    """If `dest` is a landing-page HTML linking to a CDS file, fetch that file
    (one hop) and parse it. Returns (values, file_year) or ({}, None)."""
    try:
        with open(dest, "r", encoding="utf-8", errors="ignore") as fh:
            html = fh.read()
    except Exception:
        return {}, None
    link = _find_cds_link(html, suburl)
    if not link or link == suburl:
        return {}, None
    d2 = os.path.join(CACHE, f"{unitid}_{idx}_cds.bin")
    kind, _ = fetch(link, d2)
    if kind in ("blocked", "error"):
        return {}, None
    try:
        return _parse_dest(d2, link)
    except Exception:
        return {}, None


def process_one(unitid, name, url):
    """Fetch + parse a CDS for one school; write to the appropriate checkpoint CSV.
    Returns (outcome, detail)."""
    row, reason = _attempt(unitid, name, url)
    if row is None:
        write_blocked(unitid, name, url, reason)
        return "blocked", reason
    row["unitid"] = unitid
    row["name"] = name
    row["cds_url"] = url
    write_data(row)
    nvals = sum(1 for f in DATA_FIELDS[4:] if row[f] != "")
    return "data", f"{nvals} values, {row['cds_year']}"


def run_worklist(path, limit=None):
    """Fetch+parse every school in a worklist CSV (unitid,url[,name]) and write
    to the current OUT_DATA / OUT_BLOCKED (usually per-worker part files).

    Resumable: skips unitids already present in the two output files. Rows with a
    blank url are logged blocked/not_found (the discovery step found no CDS).
    Checkpointed after every school by process_one's append.
    """
    done = _load_done()
    processed = 0
    with open(path, newline="") as fh:
        rows = list(csv.DictReader(fh))
    for r in rows:
        if not r.get("unitid"):
            continue
        uid = int(r["unitid"])
        if uid in done:
            continue
        name = r.get("name", "")
        url = (r.get("url") or "").strip()
        if not url:
            write_blocked(uid, name, "", "not_found")
            done.add(uid)
            processed += 1
            print(f"[{uid}] {name[:40]:40s} -> blocked: not_found", flush=True)
            continue
        outcome, detail = process_one(uid, name, url)
        done.add(uid)
        processed += 1
        print(f"[{uid}] {name[:40]:40s} -> {outcome}: {detail}", flush=True)
        if limit and processed >= limit:
            break
    print(f"\nWorklist done: {processed} processed this run.")


def load_discovered():
    urls = {}
    if not os.path.exists(DISCOVERED):
        return urls
    with open(DISCOVERED, newline="") as fh:
        for r in csv.DictReader(fh):
            if r.get("unitid") and r.get("url"):
                urls[int(r["unitid"])] = r["url"].strip()
    return urls


def run_batch(limit=None):
    schools = load_national_schools()
    done = _load_done()
    urls = load_discovered()
    processed = 0
    for s in schools:
        uid = s["unitid"]
        if uid in done:
            continue
        if uid not in urls:
            continue  # no discovered URL yet (agent supplies these)
        outcome, detail = process_one(uid, s["name"], urls[uid])
        processed += 1
        print(f"[{uid}] {s['name'][:40]:40s} -> {outcome}: {detail}", flush=True)
        if limit and processed >= limit:
            break
    print(f"\nBatch done: {processed} processed this run.")


# ---------------------------------------------------------------------------
# Re-run the blocked roster (recover with the hardened parser)
# ---------------------------------------------------------------------------
@contextlib.contextmanager
def _time_limit(seconds):
    """Hard wall-clock cap for one school so a pathological file can't stall the
    batch. SIGALRM only fires on the main thread (where rerun_blocked runs)."""
    def _handler(signum, frame):
        raise TimeoutError()
    old = signal.signal(signal.SIGALRM, _handler)
    signal.alarm(int(seconds))
    try:
        yield
    finally:
        signal.alarm(0)
        signal.signal(signal.SIGALRM, old)


def _read_csv(path):
    if not os.path.exists(path):
        return []
    with open(path, newline="") as fh:
        return list(csv.DictReader(fh))


def _write_all(path, fields, rows):
    tmp = path + ".tmp"
    with open(tmp, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fields})
    os.replace(tmp, path)


def _n_values(row):
    """How many of the extractable fields (DATA_FIELDS[4:]) this row fills."""
    return sum(1 for f in DATA_FIELDS[4:] if row.get(f, "") not in ("", None))


def merge_parts(parts_dir=None):
    """Merge per-worker part files into cds_data.csv / cds_blocked.csv.

    Reads data/_parts/data_*.csv and data/_parts/blocked_*.csv (plus any existing
    main CSVs), deduping by unitid: a school with an extracted row always beats a
    blocked row, and among data rows the one filling the most fields wins. Also
    folds any data/_parts/urls_*.csv into cds_discovered_urls.csv. Idempotent.
    """
    import glob as _glob
    parts_dir = parts_dir or os.path.join(DATA, "_parts")

    data_by_uid = {}
    for r in _read_csv(OUT_DATA):
        if r.get("unitid"):
            data_by_uid[int(r["unitid"])] = r
    for pf in sorted(_glob.glob(os.path.join(parts_dir, "data_*.csv"))):
        for r in _read_csv(pf):
            if not r.get("unitid"):
                continue
            uid = int(r["unitid"])
            if uid not in data_by_uid or _n_values(r) > _n_values(data_by_uid[uid]):
                data_by_uid[uid] = r

    blocked_by_uid = {}
    for r in _read_csv(OUT_BLOCKED):
        if r.get("unitid"):
            blocked_by_uid[int(r["unitid"])] = r
    for pf in sorted(_glob.glob(os.path.join(parts_dir, "blocked_*.csv"))):
        for r in _read_csv(pf):
            if r.get("unitid"):
                blocked_by_uid[int(r["unitid"])] = r
    # data wins: drop any blocked school that was extracted somewhere.
    for uid in list(blocked_by_uid):
        if uid in data_by_uid:
            del blocked_by_uid[uid]

    _write_all(OUT_DATA, DATA_FIELDS, list(data_by_uid.values()))
    _write_all(OUT_BLOCKED, BLOCKED_FIELDS, list(blocked_by_uid.values()))

    # Fold discovered URLs (unitid,url,name) into cds_discovered_urls.csv.
    url_rows = {}
    for r in _read_csv(DISCOVERED):
        if r.get("unitid") and r.get("url"):
            url_rows[int(r["unitid"])] = r
    for pf in sorted(_glob.glob(os.path.join(parts_dir, "urls_*.csv"))):
        for r in _read_csv(pf):
            if r.get("unitid") and (r.get("url") or "").strip():
                url_rows[int(r["unitid"])] = {
                    "unitid": r["unitid"], "url": r["url"].strip(),
                    "name": r.get("name", "")}
    if url_rows:
        _write_all(DISCOVERED, ["unitid", "url", "name"], list(url_rows.values()))

    print(f"Merged: {len(data_by_uid)} in cds_data.csv, "
          f"{len(blocked_by_uid)} in cds_blocked.csv, "
          f"{len(url_rows)} discovered URLs.")


def backfill(limit=None):
    """Re-parse every row already in cds_data.csv (cache first, else its known
    cds_url) and fill any BLANK field — in particular the newly-added
    pct_demonstrated_need. Existing non-blank values are never overwritten, so
    the pass is idempotent and safe to re-run. Rewrites cds_data.csv (adding the
    new column) after every school, so it is fully resumable.
    """
    rows = _read_csv(OUT_DATA)
    filled = 0
    for i, r in enumerate(rows):
        # already complete for the new field? skip the re-parse.
        if r.get("pct_demonstrated_need", "") not in ("", None):
            continue
        uid = int(r["unitid"])
        url = r.get("cds_url", "")
        try:
            with _time_limit(90):
                fresh, _ = _attempt(uid, r.get("name", ""), url, use_cache=True)
        except TimeoutError:
            fresh = None
        if fresh:
            added = []
            for f in DATA_FIELDS[4:]:
                if r.get(f, "") in ("", None) and fresh.get(f, "") not in ("", None):
                    r[f] = fresh[f]
                    added.append(f)
            if added:
                filled += 1
                print(f"[{uid}] {r.get('name','')[:34]:34s} +{','.join(added)}",
                      flush=True)
        _write_all(OUT_DATA, DATA_FIELDS, rows)
        if limit and (i + 1) >= limit:
            break
    ndn = sum(1 for r in rows if r.get("pct_demonstrated_need", "") not in ("", None))
    print(f"\nBackfill done: {filled} rows gained values; "
          f"{ndn}/{len(rows)} now have pct_demonstrated_need.")


def rerun_blocked(limit=None):
    """Re-attempt every school in cds_blocked.csv with the hardened parser.

    Recovered rows move blocked -> cds_data.csv (dedupe by unitid, data wins);
    genuinely-unreachable rows stay in cds_blocked.csv for the browser tier.
    Both CSVs are rewritten after every school, so the pass is fully resumable:
    a recovered school leaves the blocked roster and is skipped on re-run.
    """
    data_by_uid = {}
    for r in _read_csv(OUT_DATA):
        if r.get("unitid"):
            data_by_uid[int(r["unitid"])] = r
    blocked = _read_csv(OUT_BLOCKED)

    remaining = []
    recovered = 0
    for i, br in enumerate(blocked):
        uid = int(br["unitid"])
        name = br.get("name", "")
        url = br.get("attempted_url", "")
        if uid in data_by_uid:
            continue  # already extracted in a prior pass
        try:
            with _time_limit(90):
                row, reason = _attempt(uid, name, url, use_cache=True)
        except TimeoutError:
            row, reason = None, "parse_timeout"
        if row is not None:
            row["unitid"] = uid
            row["name"] = name
            row["cds_url"] = url
            data_by_uid[uid] = row
            recovered += 1
            nvals = sum(1 for f in DATA_FIELDS[4:] if row[f] != "")
            print(f"[{uid}] RECOVER {name[:38]:38s} {nvals} vals, {row['cds_year']}",
                  flush=True)
        else:
            br["reason"] = reason
            remaining.append(br)
            print(f"[{uid}] blocked {name[:38]:38s} {reason}", flush=True)

        # checkpoint: data (recoveries) + complete blocked roster (processed
        # survivors + not-yet-processed tail).
        _write_all(OUT_DATA, DATA_FIELDS, list(data_by_uid.values()))
        _write_all(OUT_BLOCKED, BLOCKED_FIELDS, remaining + blocked[i + 1:])
        if limit and (i + 1) >= limit:
            break

    print(f"\nRerun done: recovered {recovered}, "
          f"{len(data_by_uid)} in cds_data.csv, {len(remaining)} still blocked.")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main(argv=None):
    ap = argparse.ArgumentParser(description="CDS free-tier extraction pipeline")
    ap.add_argument("--batch", action="store_true",
                    help="process national schools from cds_discovered_urls.csv")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--one", action="store_true", help="process a single URL")
    ap.add_argument("--unitid", type=int)
    ap.add_argument("--name", default="")
    ap.add_argument("--url", default="")
    ap.add_argument("--parse-file", help="parse a local PDF/HTML/XLSX file (no fetch)")
    ap.add_argument("--worklist", help="CSV of unitid,url[,name] to fetch+parse")
    ap.add_argument("--data-out", help="override cds_data.csv output path")
    ap.add_argument("--blocked-out", help="override cds_blocked.csv output path")
    ap.add_argument("--rerun-blocked", action="store_true",
                    help="re-attempt cds_blocked.csv with the hardened parser")
    ap.add_argument("--backfill", action="store_true",
                    help="fill blank fields (e.g. pct_demonstrated_need) on "
                         "existing cds_data.csv rows from cache/known URLs")
    ap.add_argument("--merge-parts", action="store_true",
                    help="merge data/_parts/*.csv into the main CSVs (data wins)")
    args = ap.parse_args(argv)

    # Per-worker output overrides (used by the sharded discovery run).
    global OUT_DATA, OUT_BLOCKED
    if args.data_out:
        OUT_DATA = args.data_out
    if args.blocked_out:
        OUT_BLOCKED = args.blocked_out

    if args.worklist:
        run_worklist(args.worklist, limit=args.limit)
        return

    if args.parse_file:
        path = args.parse_file
        low = path.lower()
        if low.endswith(".pdf"):
            values, fy = parse_pdf_file(path, args.url)
            row, status, reason = _assemble_row([(values, fy)]) if values \
                else (None, "parse_fail", "pdf_parse_fail")
        elif low.endswith(".xlsx"):
            values, fy = xlsx_values(path, args.url)
            row, status, reason = _assemble_row([(values, fy)]) if values \
                else (None, "parse_fail", "pdf_parse_fail")
        else:
            row, status, reason = parse_cds_text(html_to_text(path), args.url)
        print(json.dumps({"status": status, "reason": reason, "row": row}, indent=2))
        return

    if args.merge_parts:
        merge_parts()
        return

    if args.backfill:
        backfill(limit=args.limit)
        return

    if args.rerun_blocked:
        rerun_blocked(limit=args.limit)
        return

    if args.one:
        if not (args.unitid and args.url):
            ap.error("--one requires --unitid and --url")
        outcome, detail = process_one(args.unitid, args.name, args.url)
        print(f"{outcome}: {detail}")
        return

    if args.batch:
        run_batch(limit=args.limit)
        return

    ap.print_help()


if __name__ == "__main__":
    main()
